"""Build a human-reviewable mapping workbook.

This creates a CSV showing every candidate destination row in Tier 1/Tier 2.
Joe can review it and we can convert it into the final JSON mapping.

Run:
    python build_mapping_workbook.py --tier1 "C:/.../Tier 1 Daily.xlsx" --tier2 "C:/.../Tier 2 Weekly.xlsx" --out mapping_candidates.csv
"""

from __future__ import annotations

import argparse
import csv
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SECTION_HINTS = [
    "HOURS WORKED",
    "LAY % BY HOUSE",
    "MORTALITY BY HOUSE",
    "EGG QUALITY",
    "FEED CONSUMED",
    "MILL",
    "WATER BY HOUSE",
    "TEMPERATURE",
    "THROUGHPUT",
    "QUALITY",
    "RUN RATE",
    "STOPS",
    "DOWNTIME",
    "OEE",
]


def text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def find_current_section(value: str, current: str) -> str:
    upper = value.upper()
    for hint in SECTION_HINTS:
        if hint in upper:
            return value
    return current


def scan_workbook(path: Path, workbook_name: str) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=False, read_only=True)
    rows: list[dict[str, Any]] = []
    try:
        for ws in wb.worksheets:
            section = ""
            for r in range(1, ws.max_row + 1):
                first = text(ws.cell(row=r, column=1).value)
                if first:
                    section = find_current_section(first, section)

                # Candidate metric rows have a label in col A and daily columns across B-H or B-I.
                if not first:
                    continue
                if first.lower() in {"house", "metric", "department", "barn labor"}:
                    continue
                daily_values = [ws.cell(row=r, column=c).value for c in range(2, min(ws.max_column, 9) + 1)]
                if any(v is not None for v in daily_values):
                    rows.append({
                        "workbook": workbook_name,
                        "sheet": ws.title,
                        "section": section,
                        "metric_label": first,
                        "row": r,
                        "candidate_today_cell": f"H{r}",
                        "notes": "Verify source file, source sheet, and source cell/formula",
                    })
        return rows
    finally:
        wb.close()


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--tier1", required=True)
    parser.add_argument("--tier2", required=True)
    parser.add_argument("--out", default="mapping_candidates.csv")
    args = parser.parse_args()

    rows = []
    rows.extend(scan_workbook(Path(args.tier1), "Tier 1 Daily.xlsx"))
    rows.extend(scan_workbook(Path(args.tier2), "Tier 2 Weekly.xlsx"))

    with Path(args.out).open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["workbook", "sheet", "section", "metric_label", "row", "candidate_today_cell", "notes"],
        )
        writer.writeheader()
        writer.writerows(rows)

    print(f"Wrote {len(rows)} mapping candidate rows to {args.out}")


if __name__ == "__main__":
    main()

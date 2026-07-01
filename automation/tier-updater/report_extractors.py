"""Source report extractors for Rushtown Tier Automation.

These extractors are intentionally label-based where possible so the automation
is less brittle than fixed cell references.
"""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, Optional

from openpyxl import load_workbook

from excel_helpers import normalize_text


@dataclass
class DepartmentHours:
    by_location_department: Dict[tuple[str, str], float]
    by_department: Dict[str, float]
    report_total: float


def extract_hours_by_department(path: Path) -> DepartmentHours:
    """Read an Hours_by_department report.

    Expected columns include Location Title, Department Title, and Hours.
    Works even when the column order changes.
    """
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        header_row = None
        headers: dict[str, int] = {}
        for row in ws.iter_rows(min_row=1, max_row=min(20, ws.max_row)):
            values = [normalize_text(c.value) for c in row]
            if "location title" in values and "department title" in values and "hours" in values:
                header_row = row[0].row
                headers = {normalize_text(c.value): c.column for c in row if c.value is not None}
                break
        if header_row is None:
            raise ValueError(f"Could not find header row in {path}")

        by_loc_dept: dict[tuple[str, str], float] = defaultdict(float)
        by_dept: dict[str, float] = defaultdict(float)
        total = 0.0

        loc_col = headers["location title"]
        dept_col = headers["department title"]
        hours_col = headers["hours"]

        for row_idx in range(header_row + 1, ws.max_row + 1):
            loc = ws.cell(row=row_idx, column=loc_col).value
            dept = ws.cell(row=row_idx, column=dept_col).value
            hours = ws.cell(row=row_idx, column=hours_col).value
            name = ws.cell(row=row_idx, column=1).value
            if normalize_text(name).startswith("report total"):
                try:
                    total = float(hours or 0)
                except Exception:
                    pass
                continue
            if loc is None or dept is None or hours is None:
                continue
            try:
                h = float(hours)
            except Exception:
                continue
            loc_s = str(loc).strip()
            dept_s = str(dept).strip()
            by_loc_dept[(loc_s, dept_s)] += h
            by_dept[dept_s] += h
            total += h if total == 0 else 0

        return DepartmentHours(dict(by_loc_dept), dict(by_dept), total)
    finally:
        wb.close()


def extract_first_numeric_by_label(path: Path, possible_labels: Iterable[str]) -> Optional[float]:
    """Find first matching label and return the first numeric value to the right."""
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                row_values = [normalize_text(c.value) for c in row]
                for label in possible_labels:
                    target = normalize_text(label)
                    if target in row_values:
                        idx = row_values.index(target)
                        for cell in row[idx + 1:]:
                            try:
                                if cell.value is not None:
                                    return float(cell.value)
                            except Exception:
                                continue
        return None
    finally:
        wb.close()


def extract_farm_record_snapshot(path: Path) -> dict[str, Any]:
    """Return a basic snapshot from a daily farm production workbook.

    The exact source workbook formats can vary, so this returns a dictionary of
    detected label values. The final mapping will use this plus verified cells.
    """
    labels = {
        "lay_percent": ["lay %", "lay percent", "% lay", "production %"],
        "mortality": ["mortality", "mort", "deads", "dead birds"],
        "feed": ["feed consumed", "feed", "feed lbs", "lbs feed"],
        "water": ["water", "water consumed", "gallons"],
        "birds": ["birds", "bird count", "current birds"],
        "high_temp": ["high temp", "hi temp", "house high"],
        "low_temp": ["low temp", "lo temp", "house low"],
    }
    return {key: extract_first_numeric_by_label(path, vals) for key, vals in labels.items()}

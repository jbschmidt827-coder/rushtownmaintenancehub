"""Excel helper functions for Rushtown Tier Automation."""

from __future__ import annotations

from copy import copy
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lower().replace("\n", " ").replace("  ", " ")


def find_label_cell(ws: Worksheet, label: str, max_row: int | None = None, max_col: int | None = None):
    target = normalize_text(label)
    for row in ws.iter_rows(max_row=max_row or ws.max_row, max_col=max_col or ws.max_column):
        for cell in row:
            if normalize_text(cell.value) == target:
                return cell
    return None


def find_cell_contains(ws: Worksheet, text: str, max_row: int | None = None, max_col: int | None = None):
    target = normalize_text(text)
    for row in ws.iter_rows(max_row=max_row or ws.max_row, max_col=max_col or ws.max_column):
        for cell in row:
            if target in normalize_text(cell.value):
                return cell
    return None


def find_date_column(ws: Worksheet, run_date: date, header_rows: Iterable[int] = range(1, 40)) -> Optional[int]:
    candidates = {
        run_date.strftime("%a %m/%d").lower(),
        run_date.strftime("%m/%d").lower(),
        run_date.strftime("%-m/%-d").lower() if hasattr(run_date, "strftime") else "",
    }
    for row_number in header_rows:
        for cell in ws[row_number]:
            value = cell.value
            if isinstance(value, datetime):
                if value.date() == run_date:
                    return cell.column
            if isinstance(value, date):
                if value == run_date:
                    return cell.column
            text = normalize_text(value)
            if text in candidates or run_date.strftime("%m/%d") in text:
                return cell.column
    return None


def copy_style(source_cell, destination_cell) -> None:
    if source_cell.has_style:
        destination_cell._style = copy(source_cell._style)
    destination_cell.number_format = source_cell.number_format
    destination_cell.alignment = copy(source_cell.alignment)
    destination_cell.border = copy(source_cell.border)
    destination_cell.fill = copy(source_cell.fill)
    destination_cell.font = copy(source_cell.font)
    destination_cell.protection = copy(source_cell.protection)


def copy_left_style(ws: Worksheet, row: int, col: int) -> None:
    if col <= 1:
        return
    copy_style(ws.cell(row=row, column=col - 1), ws.cell(row=row, column=col))


def read_workbook_value(path: Path, sheet_name: str, cell: str) -> Any:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        return wb[sheet_name][cell].value
    finally:
        wb.close()


def get_value_by_label(path: Path, sheet_name: str, label: str, offset_rows: int = 0, offset_cols: int = 1) -> Any:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[sheet_name]
        label_cell = find_label_cell(ws, label)
        if label_cell is None:
            raise ValueError(f"Label not found in {path.name} {sheet_name}: {label}")
        return ws.cell(row=label_cell.row + offset_rows, column=label_cell.column + offset_cols).value
    finally:
        wb.close()

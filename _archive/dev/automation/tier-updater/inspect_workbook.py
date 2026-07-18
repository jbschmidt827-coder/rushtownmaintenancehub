"""Workbook inspection helper for Rushtown Tier Automation.

This tool creates a CSV map of workbook sheets, cell values, formulas, merged ranges,
and conditional formatting ranges. It does not change the workbook.

Run:
    python inspect_workbook.py "C:/path/to/Tier 1 Daily.xlsx" --out tier1_map.csv
"""

from __future__ import annotations

import argparse
import csv
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


def iter_used_cells(ws) -> Iterable[dict]:
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            yield {
                "sheet": ws.title,
                "cell": cell.coordinate,
                "row": cell.row,
                "column": cell.column,
                "value": str(cell.value),
                "is_formula": str(isinstance(cell.value, str) and cell.value.startswith("=")),
                "number_format": cell.number_format,
                "style_id": cell.style_id,
            }


def inspect_workbook(workbook_path: Path, out_path: Path) -> None:
    wb = load_workbook(workbook_path, data_only=False)
    try:
        rows = []
        for ws in wb.worksheets:
            rows.extend(iter_used_cells(ws))

        out_path.parent.mkdir(parents=True, exist_ok=True)
        with out_path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(
                f,
                fieldnames=["sheet", "cell", "row", "column", "value", "is_formula", "number_format", "style_id"],
            )
            writer.writeheader()
            writer.writerows(rows)

        print(f"Wrote workbook map: {out_path}")

        print("Sheets:")
        for ws in wb.worksheets:
            print(f"- {ws.title}: {ws.max_row} rows x {ws.max_column} columns")
            print(f"  Merged ranges: {len(list(ws.merged_cells.ranges))}")
            print(f"  Conditional formatting ranges: {len(ws.conditional_formatting)}")
    finally:
        wb.close()


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", help="Workbook path")
    parser.add_argument("--out", default="workbook_map.csv", help="Output CSV path")
    args = parser.parse_args()
    inspect_workbook(Path(args.workbook), Path(args.out))


if __name__ == "__main__":
    main()

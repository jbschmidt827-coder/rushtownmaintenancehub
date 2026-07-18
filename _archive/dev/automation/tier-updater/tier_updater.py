"""Rushtown Tier Updater

Updates existing Tier 1 and Tier 2 Excel workbooks from daily source reports.

This starter is intentionally safe:
- It backs up the live workbooks before changing anything.
- It edits existing workbooks instead of rebuilding them, preserving layout and formatting.
- It supports a mapping file so cells can be verified one by one.

Run:
    python tier_updater.py --config config.json --date today
"""

from __future__ import annotations

import argparse
import json
import logging
import shutil
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


@dataclass(frozen=True)
class AppConfig:
    tier_meetings_folder: Path
    daily_reports_folder: Path
    backup_folder: Path
    log_folder: Path
    tier_1_file: str
    tier_2_file: str
    mapping_file: Path
    fail_on_missing_required_report: bool = False


def load_config(path: Path) -> AppConfig:
    raw = json.loads(path.read_text(encoding="utf-8"))
    base = path.parent
    mapping = Path(raw["mapping_file"])
    if not mapping.is_absolute():
        mapping = base / mapping
    return AppConfig(
        tier_meetings_folder=Path(raw["tier_meetings_folder"]),
        daily_reports_folder=Path(raw["daily_reports_folder"]),
        backup_folder=Path(raw["backup_folder"]),
        log_folder=Path(raw["log_folder"]),
        tier_1_file=raw["tier_1_file"],
        tier_2_file=raw["tier_2_file"],
        mapping_file=mapping,
        fail_on_missing_required_report=bool(raw.get("fail_on_missing_required_report", False)),
    )


def resolve_run_date(value: str) -> date:
    if value.lower() == "today":
        return date.today()
    if value.lower() == "yesterday":
        from datetime import timedelta
        return date.today() - timedelta(days=1)
    return datetime.strptime(value, "%Y-%m-%d").date()


def setup_logging(config: AppConfig, run_date: date) -> None:
    config.log_folder.mkdir(parents=True, exist_ok=True)
    log_path = config.log_folder / f"tier_update_{run_date:%Y%m%d}.log"
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
        handlers=[logging.FileHandler(log_path, encoding="utf-8"), logging.StreamHandler()],
    )


def backup_file(path: Path, backup_folder: Path, run_date: date) -> Path:
    if not path.exists():
        raise FileNotFoundError(f"Tier workbook not found: {path}")
    backup_folder.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = backup_folder / f"{path.stem}_{run_date:%Y%m%d}_{stamp}{path.suffix}"
    shutil.copy2(path, backup_path)
    logging.info("Backup created: %s", backup_path)
    return backup_path


def load_mapping(mapping_path: Path) -> Dict[str, Any]:
    if not mapping_path.exists():
        raise FileNotFoundError(f"Mapping file not found: {mapping_path}")
    return json.loads(mapping_path.read_text(encoding="utf-8"))


def find_latest_report(folder: Path, patterns: Iterable[str]) -> Optional[Path]:
    matches = []
    for pattern in patterns:
        matches.extend(folder.glob(pattern))
    matches = [p for p in matches if p.is_file() and not p.name.startswith("~$")]
    if not matches:
        return None
    return max(matches, key=lambda p: p.stat().st_mtime)


def read_source_cell(report_path: Path, sheet_name: str, cell: str) -> Any:
    wb = load_workbook(report_path, data_only=True, read_only=True)
    try:
        ws = wb[sheet_name]
        return ws[cell].value
    finally:
        wb.close()


def copy_left_style(ws: Worksheet, cell_address: str) -> None:
    """Copy style from the cell immediately to the left, if available.

    This is the safe default for date columns: the new daily column should behave
    like the previous day. Existing conditional formatting rules are preserved by
    editing the workbook in place.
    """
    from copy import copy
    cell = ws[cell_address]
    if cell.column <= 1:
        return
    source = ws.cell(row=cell.row, column=cell.column - 1)
    if source.has_style:
        cell._style = copy(source._style)
    if source.number_format:
        cell.number_format = source.number_format
    if source.alignment:
        cell.alignment = copy(source.alignment)
    if source.border:
        cell.border = copy(source.border)


def apply_mapping_to_workbook(workbook_path: Path, mapping: Dict[str, Any], config: AppConfig) -> None:
    wb = load_workbook(workbook_path)
    try:
        workbook_key = workbook_path.name
        rules = mapping.get("workbooks", {}).get(workbook_key, [])
        logging.info("Applying %s mapped rules to %s", len(rules), workbook_key)

        for rule in rules:
            enabled = rule.get("enabled", True)
            if not enabled:
                continue

            dest_sheet = rule["destination"]["sheet"]
            dest_cell = rule["destination"]["cell"]
            source = rule["source"]
            report = find_latest_report(config.daily_reports_folder, source["file_patterns"])
            required = bool(rule.get("required", False))

            if report is None:
                message = f"Missing report for {rule.get('name', dest_cell)} using patterns {source['file_patterns']}"
                if required and config.fail_on_missing_required_report:
                    raise FileNotFoundError(message)
                logging.warning(message)
                continue

            value = read_source_cell(report, source["sheet"], source["cell"])
            ws = wb[dest_sheet]
            copy_left_style(ws, dest_cell)
            ws[dest_cell] = value
            logging.info("%s!%s <- %s from %s!%s", dest_sheet, dest_cell, value, report.name, source["cell"])

        wb.save(workbook_path)
        logging.info("Saved updated workbook: %s", workbook_path)
    finally:
        wb.close()


def run(config_path: Path, run_date: date) -> None:
    config = load_config(config_path)
    setup_logging(config, run_date)
    logging.info("Starting Rushtown Tier update for %s", run_date)

    mapping = load_mapping(config.mapping_file)
    tier_1_path = config.tier_meetings_folder / config.tier_1_file
    tier_2_path = config.tier_meetings_folder / config.tier_2_file

    backup_file(tier_1_path, config.backup_folder, run_date)
    backup_file(tier_2_path, config.backup_folder, run_date)

    apply_mapping_to_workbook(tier_1_path, mapping, config)
    apply_mapping_to_workbook(tier_2_path, mapping, config)

    logging.info("Rushtown Tier update completed")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--config", default="config.json", help="Path to config.json")
    parser.add_argument("--date", default="today", help="today, yesterday, or YYYY-MM-DD")
    args = parser.parse_args()
    run(Path(args.config), resolve_run_date(args.date))


if __name__ == "__main__":
    main()
